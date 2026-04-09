from __future__ import annotations

from typing import Iterator

from openpyxl import load_workbook

from elbench.loaders.base import BaseLoader
from elbench.loaders.normalizer import RecordNormalizer
from elbench.registry import ResolvedRegistryItem
from elbench.schemas.evaluation import RawRecord, Sample


class XlsxLoader(BaseLoader):
    def __init__(self) -> None:
        self._normalizer = RecordNormalizer()

    def iter_raw_records(self, item: ResolvedRegistryItem) -> Iterator[RawRecord]:
        workbook = load_workbook(item.path, read_only=True, data_only=True)
        sheet_name = item.mapping.sheet_name or workbook.sheetnames[0]
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
        for row_index, row_values in enumerate(rows, start=2):
            record = {
                header: value
                for header, value in zip(headers, row_values)
                if header
            }
            if not any(value not in (None, "") for value in record.values()):
                continue
            yield RawRecord(
                source_path=str(item.path),
                source_file=item.entry.canonical_name,
                row_index=row_index,
                record=record,
            )
        workbook.close()

    def iter_samples(self, item: ResolvedRegistryItem) -> Iterator[Sample]:
        for raw_record in self.iter_raw_records(item):
            yield self._normalizer.normalize(item, raw_record)

